#!/usr/bin/python
# coding=utf-8
from openpyxl import load_workbook
import cx_Oracle
import os
import collections
os.environ['NLS_LANG'] = 'SIMPLIFIED CHINESE_CHINA.UTF8'

class GeneralBusinessKeywordLibrary(object):
    def __init__(self):
        self.conn = None
        self.cursor = None
        pass

    '''
    功能：组装json消息，适用于创建活动接口的消息组装，已使用的接口有：单主品满赠阶梯、单主品满赠每满、双主品满赠组合、单主品满赠、双主品满赠阶梯组合、双主品满赠每满组合、单主品满赠累单、单主品满赠阶梯累单、单主品满赠每满累单
    编写人：刘义颖
    输入参数：excel_path Excel文件路径
              sheet_name sheet名称
              column_number 列号
    返回结果：True/False
    日期： 2019-08-22
    修改记录：
    '''
    def combine_json_create_activity(self, excel_path, sheet_name, column_number):
        wb = load_workbook(excel_path)
        sheet = wb[sheet_name]
        max_rows = sheet.max_row
        massage_string = '{'
        col_1_before = '1.1.1.1.1'
        # 数据模板从第三行开始
        for i in range(3, max_rows+1):
            # 数据列
            col_data = sheet.cell(row=i, column=int(column_number)).value
            # 第一列：层级编号
            col_1 = str(sheet.cell(row=i, column=1).value)
            # 下一行层级编号
            col_1_next = str(sheet.cell(row=i+1, column=1).value)
            # 第二列：字段名称
            col_2 = sheet.cell(row=i, column=2).value
            # 第三列，字段类型
            col_3 = sheet.cell(row=i, column=3).value
            # 下一层字段值
            col_data_next = sheet.cell(row=i + 1, column=int(column_number)).value
            # 再下一层字段值
            col_data_next_next = sheet.cell(row=i + 2, column=int(column_number)).value
            # 如果数据列标记为NULL，则认为该字段不需要，跳过
            if col_data == 'NULL':
                if col_1_next == 'None':
                    if len(col_1) == 1:
                        massage_string = massage_string + '}'
                    elif len(col_1) == 7:
                        massage_string = massage_string + '}]}]}]}'
                    elif len(col_1) == 5:
                        massage_string = massage_string + '}]}]}'
                    else:
                        massage_string = massage_string + '}]}'
                elif str(col_2) == 'presentSource':
                    massage_string = massage_string + '}]},{'
                else:
                    col_1_before = col_1
                continue
            # 开始字段判断，如果这一行是某一层级的第一个字段，需要有大括号
            if len(col_1_before) < len(col_1):
                massage_string = massage_string + '{'
            # 存放该行数据
            if col_3 == 'Array':
                massage_string = massage_string + '"'+col_2+'"'+':['
            elif col_3 == 'Map':
                massage_string = massage_string + '"' + col_2 + '"' + ':{'
            elif col_3 == 'Object':
                massage_string = massage_string + '"' + col_2 + '"' + ':{'
            elif col_3 == 'String':
                massage_string = massage_string + '"' + col_2 + '":"' + str(col_data)+'"'
            else:
                massage_string = massage_string + '"' + col_2 + '":' + str(col_data)
            # 结束字段判断，如果下一行和当前行层级编号一样，说明后面有同级字段，增加逗号即可
            if col_1_next == 'None':
                if len(col_1) == 3:
                    massage_string = massage_string + '}]}'
                elif len(col_1) == 1:
                    massage_string = massage_string + '}'
                elif len(col_1) == 5:
                    massage_string = massage_string + '}]}]}'
                else:
                    massage_string = massage_string + '}]}]}]}'
            elif len(col_1) == len(col_1_next):
                if col_data_next != 'NULL':
                    if col_1 == col_1_next:
                        massage_string = massage_string + ','
                    else:
                        massage_string = massage_string + '},{'
                else:
                    flag = 0
                    for j in range(i + 1, max_rows + 1):
                        col_after = sheet.cell(row=j, column=int(column_number)).value
                        col_number = str(sheet.cell(row=j, column=1).value)
                        if len(col_1) != len(col_number):
                            break
                        if col_after != 'NULL' and col_after != 'None' and len(col_1) == len(col_number):
                            flag = 1
                    if flag:
                        if col_1 == col_1_next:
                            massage_string = massage_string + ','
                        else:
                            massage_string = massage_string + '},{'
            elif len(col_1) > len(col_1_next):
                if str(col_2) == 'presentSource':
                    if col_1 == '1.1.2':
                        massage_string = massage_string + '}]}],'
                    elif len(col_1) == '7' and len(col_1_next) == '5':
                        massage_string = massage_string + '},{'
                    #该分支仅适用于单主品满赠-双规则的活动创建
                    elif col_1 == '1.1.1.1' and col_1_next == '1.2':
                         massage_string = massage_string + '}]}]},{'
                    else:
                        massage_string = massage_string + '}]},{'
                elif str(col_2) == 'PDCQuota':
                    # if col_1 == '1.1.2':
                    massage_string = massage_string + '}]}],'
                    # else:
                    #     massage_string = massage_string + '}]},{'
                else:
                        massage_string = massage_string + '}],'

            col_1_before = col_1

        return massage_string

    '''
    功能：组装json消息，适用于除了创建活动之外的接口的消息组装，已使用的接口有：预计算、计算冻结、根据处理号预算补偿
    编写人：刘义颖
    输入参数：excel_path Excel文件路径
              sheet_name sheet名称
              column_number 列号
    返回结果：True/False
    日期： 2019-08-22
    修改记录：
    '''
    def combine_json_others(self, excel_path, sheet_name, column_number):
        wb = load_workbook(excel_path,data_only=True)
        sheet = wb[sheet_name]
        max_rows = sheet.max_row
        massage_string = '{'
        col_1_before = '1.1.1.1'
        # 数据模板从第三行开始
        for i in range(3, max_rows+1):
            # 数据列
            col_data = str(sheet.cell(row=i, column=int(column_number)).value)
            # 第一列：层级编号
            col_1 = str(sheet.cell(row=i, column=1).value)
            # 下一行层级编号
            col_1_next = str(sheet.cell(row=i+1, column=1).value)
            # 第二列：字段名称
            col_2 = str(sheet.cell(row=i, column=2).value)
            # 第三列，字段类型
            col_3 = str(sheet.cell(row=i, column=3).value)
            # 下一层字段值
            col_data_next = str(sheet.cell(row=i+1, column=int(column_number)).value)
            #再下一层字段值
            col_data_next_next = str(sheet.cell(row=i + 2, column=int(column_number)).value)
            # 如果数据列标记为NULL，则认为该字段不需要，跳过
            if col_data == 'NULL':
                if col_1_next == 'None':
                    if len(col_1) == 1:
                        massage_string = massage_string + '}'
                    else:
                        massage_string = massage_string + '}]}'
                elif str(col_2) == 'presentSource':
                    massage_string = massage_string + '},{'
                else:
                    col_1_before = col_1
                continue
            # 开始字段判断，如果这一行是某一层级的第一个字段，需要有大括号
            if len(col_1_before) < len(col_1):
                massage_string = massage_string + '{'
            # 存放该行数据
            if col_3 == 'Array':
                if col_2 == 'rules':
                    massage_string = massage_string + '"' + col_2 + '"' + ':["' + col_data + '"]'
                else:
                    massage_string = massage_string + '"'+col_2+'"'+':['
            elif col_3 == 'Map':
                massage_string = massage_string + '"' + col_2 + '"' + ':{'
            elif col_3 == 'String':
                massage_string = massage_string + '"' + col_2 + '":"' + str(col_data)+'"'
            elif col_3 == 'ArrayString':
                massage_string = massage_string + '"' + col_2 + '":[' + str(col_data) + ']'
            else:
                massage_string = massage_string + '"' + col_2 + '":' + str(col_data)
            # 结束字段判断，如果下一行和当前行层级编号一样，说明后面有同级字段，增加逗号即可
            if col_1_next == 'None':
                if len(col_1) == 3:
                    massage_string = massage_string +'}]}'
                elif len(col_1) == 1:
                    massage_string = massage_string + '}'
                else:
                    massage_string = massage_string + '}]}]}]}'
            elif len(col_1) == len(col_1_next):
                if col_data_next != 'NULL':
                    if col_1 == col_1_next:
                        massage_string = massage_string + ','
                    else:
                        massage_string = massage_string + '},{'
                else:
                    flag = 0
                    for j in range(i+1, max_rows + 1):
                        col_after = sheet.cell(row=j, column=int(column_number)).value
                        col_number = str(sheet.cell(row=j, column=1).value)
                        if len(col_1) != len(col_number):
                            break
                        if col_after!='NULL' and col_after!='None' and len(col_1)==len(col_number):
                            flag = 1
                    if flag:
                        massage_string = massage_string + ','

            elif len(col_1) > len(col_1_next):
                if str(col_2) == 'presentSource':
                    if col_1 == '1.1.2':
                        massage_string = massage_string + '}]}],'
                    elif col_1 == '1.1.1.1':
                        massage_string = massage_string + '}]}]},{'
                    else:
                        massage_string = massage_string + '}]},{'
                else:
                    massage_string = massage_string + '}],'
            col_1_before = col_1
        if massage_string =='{}]}':
            massage_string = '{}'
        return massage_string


    '''
    功能：数据库连接，目前仅适用于oracle数据库，后续有需要再扩展
    编写人：刘义颖
    输入参数：ip  数据库IP
              port 数据库端口
              username  数据库用户名
              password  数据库密码
              connect_type: 连接类型，支持servicename,sid两种枚举值
              connect_value:对应连接类型的值
    返回结果：True/False
    日期： 2019-08-10
    修改记录：
    '''
    def database_connection(self, ip, port, username, password, connect_type, connect_value):
        if connect_type == 'servicename':
            self.conn = cx_Oracle.connect('%s/%s@%s:%s/%s' % (username, password, ip, port, connect_value), encoding="utf8", nencoding="UTF-8")
        elif connect_type == 'sid':
            tnsname = cx_Oracle.makedsn(ip,int(port),connect_value)
            self.conn = cx_Oracle.connect(username,password,tnsname)
        else:
            return False
        return True

    def database_query(self,sql):

        os.environ['NLS_LANG'] = 'SIMPLIFIED CHINESE_CHINA.UTF8'
        self.cursor = self.conn.cursor()
        self.cursor.execute(sql)
        row = self.cursor.fetchall()
        return row

    def respect_data_convert_to_list(self, excel_path, sheet_name, col_number):
        wb = load_workbook(excel_path,data_only=True)
        sheet = wb[sheet_name]
        max_rows = sheet.max_row
        col_number = col_number.split('|')
        respect_list=[]
        # 数据模板从第2行开始
        for col_number_i in col_number:
            temp_list=[]
            for i in range(2, max_rows+1):
                temp_data = ''
                # 字段名称
                col_1 = str(sheet.cell(row=i, column=1).value)
                # 字段类型
                col_2 = str(sheet.cell(row=i, column=2).value)
                # 期望值
                col_3 = str(sheet.cell(row=i, column=int(col_number_i)).value)
                if col_3 != 'NULL':
                    if col_1 == 'message':
                        temp_data = col_3
                    else:
                        if col_2 == 'String':
                            temp_data = "'" + col_1 +"': '" + col_3 + "'"
                        else:
                            temp_data = "'" + col_1 + "': " + col_3
                    temp_list.append(temp_data)
            if temp_list != []:
                respect_list.append(temp_list)
        if respect_list == []:
            respect_list = None
        return respect_list


    def combine_data_to_string(self, excel_path, sheet_name, col_number):
        wb = load_workbook(excel_path,data_only=True)
        sheet = wb[sheet_name]
        max_rows = sheet.max_row
        result_data = '?'
        flag = 1
        # 数据模板从第3行开始
        for i in range(3, max_rows+1):
            # 字段名称
            col_1 = str(sheet.cell(row=i, column=2).value)
            # 值
            col_3 = str(sheet.cell(row=i, column=int(col_number)).value)

            if col_3 != 'NULL':
                result_data = result_data + col_1 +"=" + col_3 + "&"
        if result_data == '?':
            result_data = None
            flag = 0
        else:
            result_data = result_data[:-1]
        return result_data,flag


    '''
    功能：优化方法，组装json消息，适用于创建活动接口的消息组装
    编写人：刘义颖
    输入参数：excel_path Excel文件路径
              sheet_name sheet名称
              column_number 列号
    返回结果：True/False
    日期： 2019-08-22
    修改记录：
    '''
    def combine_json_create_activity_new(self, excel_path, sheet_name, column_number):
        wb = load_workbook(excel_path)
        sheet = wb[sheet_name]
        max_rows = sheet.max_row
        massage_string = '{'
        col_1_before = '1.1.1.1.1'
        # 数据模板从第三行开始
        for i in range(3, max_rows+1):
            # 数据列
            col_data = str(sheet.cell(row=i, column=int(column_number)).value)
            # 第一列：层级编号
            col_1 = str(sheet.cell(row=i, column=1).value)
            # 下一行层级编号
            for j in range(i + 1, max_rows + 1):
                col_after = sheet.cell(row=j, column=int(column_number)).value
                col_number = str(sheet.cell(row=j, column=1).value)
                if col_after != 'NULL':
                    col_1_next = col_number
                    break
            # 第二列：字段名称
            col_2 = sheet.cell(row=i, column=2).value
            # 第三列，字段类型
            col_3 = sheet.cell(row=i, column=3).value
            # 下一层字段名称
            for j in range(i + 1, max_rows + 1):
                col_after = sheet.cell(row=j, column=int(column_number)).value
                col_2_next = sheet.cell(row=j, column=2).value
                if col_after != 'NULL':
                    break
            # 上一层字段名称和值，为了 sku 获取
            for k in range(i-1,max_rows-1):
                col_1_befor_index = sheet.cell(row=k, column=1).value
                col_1_before_value = sheet.cell(row=k, column=2).value
                col_1_before_value_real = sheet.cell(row=k, column=int(column_number)).value

                if col_1_befor_index != 'NULL':
                    break
            #判断当前元素是否是同一层级中的最后一个元素，flag=1是同级连续还有元素
            tongji_flag = 0
            for j in range(i + 1, max_rows + 1):
                col_after = sheet.cell(row=j, column=int(column_number)).value
                col_number = str(sheet.cell(row=j, column=1).value)
                if col_1 != col_number:
                    break
                elif col_after != 'NULL' and col_after != 'None':
                    tongji_flag = 1
            # 如果数据列标记为NULL，则认为该字段不需要，跳过
            if col_data == 'NULL':
                if col_1_next == 'None':
                    massage_string = massage_string + self.fun(len(col_1))
                else:
                    col_1_before = col_1

                continue
            # bug 查阅
            # if col_1_before_value == 'presentCode':
            #     print("hello")
            # 开始字段判断，如果这一行是某一层级的第一个字段，需要有大括号
            if len(col_1_before) <len(col_1):
                massage_string = massage_string + '{'
            if len(col_1_before) ==len(col_1) and col_1_before!=col_1:
                massage_string = massage_string + '{'
            #这段代码为了解决：condition 中，sku不存在，value存在，二期： scope = ‘order’时 sku不需要填的问题 
            if len(col_1_before) ==len(col_1) and col_1_before_value=='sku' and col_1_before_value_real=='NULL':
                massage_string = massage_string + '{'
            

            # 存放该行数据
            if col_3 == 'Array':
                massage_string = massage_string + '"'+col_2+'"'+':['
            elif col_3 == 'Map':
                massage_string = massage_string + '"' + col_2 + '"' + ':{'
            elif col_3 == 'String':
                massage_string = massage_string + '"' + col_2 + '":"' + str(col_data)+'"'
            else:
                massage_string = massage_string + '"' + col_2 + '":' + str(col_data)
            # 结束字段判断，如果下一行和当前行层级编号一样，说明后面有同级字段，增加逗号即可
            if col_3 != 'Array' and col_3 != 'Map':
                #判断当前元素是否有后续字段
                quanliang_flag = 0
                for j in range(i + 1, max_rows + 1):
                    col_number_data = str(sheet.cell(row=int(j), column=int(column_number)).value)
                    if col_number_data != 'NULL':
                        quanliang_flag=1
                        break
                if quanliang_flag == 1:
                    if tongji_flag == 0:
                        if len(col_1) != len(col_1_next):
                            length = len(col_1) - len(col_1_next) + 1
                            if col_2_next == 'groupId':
                                massage_string = massage_string + self.fun(length)+'},{'
                            elif col_2_next == 'presentCode':
                                massage_string = massage_string + self.fun(length)+'},{'
                            else:
                                massage_string = massage_string + self.fun(length) + ','
                        else:
                            massage_string = massage_string + '},'
                    else:
                        massage_string = massage_string + ','
                else:
                    massage_string = massage_string + self.fun(len(col_1)+1)
            col_1_before = col_1
        return massage_string

    def fun(self,length):
        string_end = ''
        for k in range(1,int(length)):
            if k % 2 == 1:
                string_end = string_end + '}'
            else:
                string_end = string_end + ']'
        return string_end

    def getAllDirQueue(self,path):
        queue = collections.deque()
        queue.append(path)
        result = []
        while len(queue) != 0:
            dirpath = queue.popleft()
            filelist = os.listdir(dirpath)
            for listname in filelist:
                fileabspath = os.path.join(dirpath, listname)
                result.append(fileabspath)
        return result

    def oracle_database_return_dict(self,sql,ip, port, username, password, connect_type, connect_value):
        self.database_connection(ip, port, username, password, connect_type, connect_value)
        value = self.database_query(sql)
        cols = [d[0] for d in self.cursor.description]
        dic = None
        for row in value:
            dic = dict(zip(cols, row))
        self.cursor.close()
        self.conn.close()
        return dic
if __name__ == '__main__':
    input
    a = GeneralBusinessKeywordLibrary()
    # data = a.combine_json_create_activity_new("D:\\tongyong_project\\MMS_AUTO_TEST\\AutomationTest\\TestData\\单主品满减阶梯.xlsx", "创建活动接口", "7")
    # print(data)
    #data_other = a.combine_json_create_activity_new("D:\\tongyong_project\\MMS_AUTO_TEST\\AutomationTest\\TestData_2\\单主品满减.xlsx", "创建卡券", "7")

    data = a.combine_json_create_activity_new("D:\\tongyong_project\\MMS_AUTO_TEST\\AutomationTest\\TestData\\双主品满赠阶梯组合_金额.xlsx", "创建活动接口", "12")
    print(data)
    # data = a.combine_json_pre_calculate("D:\\tongyong_project\\MMS_AUTO_TEST\\AutomationTest\\TestData\\单主品满折.xlsx", "Prediction", "6")

    #data = a.combine_json_pre_calculate("C:\\Users\\86136\\Desktop\\data.xlsx", "Prediction", "6")
    # data = a.combine_json_create_activity("D:\\gitlab\\BP_PROMOTION_MC_TEST\\AutomationTest\\TestData\\双主品满赠阶梯组合.xlsx", "创建活动接口", "7")
    # data = a.combine_data_to_string("D:\\gitlab\\BP_PROMOTION_MC_TEST\\AutomationTest\\TestData\\单主品满赠.xlsx", "查询活动列表forAPP接口","8")
    # data = a.combine_json_others("D:\\gitlab\\BP_PROMOTION_MC_TEST\\AutomationTest\\TestData\\双主品满赠组合.xlsx", "预计算接口", "6")
    # data = a.respect_data_convert_to_list("D:\\gitlab\\BP_PROMOTION_MC_TEST\\AutomationTest\\TestData\\双主品满赠组合.xlsx", "预计算期望结果","3,4")
    # data = a.combine_json_create_activity("D:\\gitlab\\BP_PROMOTION_MC_TEST\\AutomationTest\\TestData\\双主品满折阶梯组合.xlsx","创建活动接口", "7")
    # print(data)
    # data = a.combine_json_create_activity_new("D:\\gitlab\\BP_PROMOTION_MC_TEST\\AutomationTest\\TestData\\双主品满折阶梯组合.xlsx",
    #                                       "创建活动接口", "7")
    # print(data)
    # a.oracle_database_return_dict("SELECT t.* FROM OWSSC.B_ALTER_BUDGET_ITEM t WHERE t.ALTER_CODE='ALT_1204282892132347906' and t.budget_type='1'",'10.203.104.131', '1531', 'OWSSC', 'cDMCN2tW', 'sid', 'QA086')
    # result = a.database_query("select * from B_ACTIVITY t where t.external_code = 'Ex100917400001'")
    # print(result)

    # data = a.combine_json_create_activity("D:\\gitlab\\OMC-TEST\\OMC-TEST\\AutomationTest\\TestData\\新建订单接口.xlsx", "测试数据", "7")
    #data = a.combine_json_create_activity("D:\\gitlab\\BP_PROMOTION_MC_TEST\\AutomationTest\\TestData\\单主品满赠-双规则.xlsx", "创建活动接口", "7")
    # data = a.combine_json_create_activity_new("D:\\gitlab\\BP_PROMOTION_MC_TEST\\AutomationTest\\TestData\\双主品满赠阶梯组合.xlsx","创建活动接口", "7")
    # print(data)
    # filepathlist = a.getAllDirQueue("D:\\gitlab\\BP_PROMOTION_MC_TEST\\AutomationTest\\TestData")
    # for item in filepathlist:
    #     data_true = a.combine_json_create_activity(item,"创建活动接口", "7")
    #     data_test = a.combine_json_create_activity_new(item,"创建活动接口", "7")
    #     if data_test != data_true:
    #         print('======================================')WSSSC/WSSSC1234
    #         print(item)
    #         print(data_true)
    #         print(data_test)
    #         print('======================================')

    #a.database_connection('10.203.104.131', '1532', 'OWOSCDV', 'AIsJLQS8', 'servicename', 'OSCDV.SGM.COM')
